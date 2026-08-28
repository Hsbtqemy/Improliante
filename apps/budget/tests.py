"""Tests du domaine budget : émission des reçus fiscaux.

Le rendu PDF (WeasyPrint) est systématiquement remplacé par un faux moteur :
les tests valident la numérotation, le snapshot et la mise en cache, pas la
mise en page PDF.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal

from apps.budget import graphiques
from apps.budget.models import (
    Adhesion,
    Categorie,
    RecuFiscal,
    Saison,
    SoldeTresorerie,
    Transaction,
)
from apps.budget.services import (
    assurer_pdf_recu,
    bilan_par_categorie,
    classeur_bilan,
    donnees_depuis_adhesion,
    emettre_recu,
    pdf_de_recu,
    tresorerie,
)
from apps.coeur.models import Membre, ParametresAssociation, Signataire, Utilisateur


def _membre(username="alice"):
    user = Utilisateur.objects.create_user(username=username, password="x")
    return Membre.objects.create(user=user)


def _emettre(**extra):
    donnees = {
        "type_versement": RecuFiscal.TypeVersement.DON,
        "montant": Decimal("50.00"),
        "date_versement": date(2026, 3, 1),
        "donateur_nom": "Jean Dupont",
    }
    donnees.update(extra)
    return emettre_recu(**donnees)


def test_emettre_recu_numerote_sans_trou(db):
    r1 = _emettre(date_emission=date(2026, 5, 1))
    r2 = _emettre(date_emission=date(2026, 6, 1))
    assert r1.numero == "R2026-0001"
    assert r2.numero == "R2026-0002"


def test_numerotation_repart_a_chaque_annee(db):
    r2026 = _emettre(date_emission=date(2026, 12, 31))
    r2027 = _emettre(date_emission=date(2027, 1, 1))
    assert r2026.numero == "R2026-0001"
    assert r2027.numero == "R2027-0001"


def test_emettre_recu_fige_le_snapshot(db):
    recu = _emettre(montant=Decimal("120.00"), donateur_nom="Marie Martin")
    assert recu.montant == Decimal("120.00")
    assert recu.donateur_nom == "Marie Martin"
    assert recu.type_versement == RecuFiscal.TypeVersement.DON


def test_donnees_depuis_adhesion(db):
    membre = _membre()
    saison = Saison.objects.create(nom="2025-2026")
    adhesion = Adhesion.objects.create(
        membre=membre,
        saison=saison,
        statut=Adhesion.Statut.PAYEE,
        montant_verse=Decimal("30.00"),
        date=date(2025, 9, 15),
    )
    donnees = donnees_depuis_adhesion(adhesion)
    assert donnees["montant"] == Decimal("30.00")
    assert donnees["type_versement"] == RecuFiscal.TypeVersement.COTISATION
    assert donnees["date_versement"] == date(2025, 9, 15)
    assert str(membre) in donnees["donateur_nom"]


def test_donnees_depuis_adhesion_adherent_sans_compte(db):
    """Un adhérent sans compte de connexion peut recevoir un reçu : le nom du
    donateur vient de la fiche (identité propre, pas d'un compte)."""
    membre = Membre.objects.create(prenom="Nina", nom="Roche")  # aucun user
    saison = Saison.objects.create(nom="2025-2026")
    adhesion = Adhesion.objects.create(
        membre=membre, saison=saison, statut=Adhesion.Statut.PAYEE, montant_verse=Decimal("20")
    )
    donnees = donnees_depuis_adhesion(adhesion)
    assert donnees["donateur_nom"] == "Nina Roche"


def test_assurer_pdf_rend_une_seule_fois(db, monkeypatch):
    appels = []

    def faux_rendu(html, *, base_url=None):
        appels.append(html)
        return b"%PDF-1.4 faux"

    monkeypatch.setattr("apps.common.pdf.html_vers_pdf", faux_rendu)

    recu = _emettre()
    assurer_pdf_recu(recu)
    assert recu.fichier  # le PDF a été créé et mis en cache
    assurer_pdf_recu(recu)  # 2e appel : le fichier existe déjà, pas de re-rendu
    assert len(appels) == 1
    assert recu.fichier.open("rb").read().startswith(b"%PDF")


def test_cerfa_utilise_le_signataire_choisi(db, monkeypatch):
    monkeypatch.setattr(
        "apps.common.pdf.html_vers_pdf", lambda html, *, base_url=None: html.encode()
    )
    sig = Signataire.objects.create(nom="Alice Martin", qualite="Présidente")
    recu = _emettre(signataire=sig)
    html = pdf_de_recu(recu).decode()
    assert "Alice Martin" in html


def test_cerfa_retombe_sur_le_signataire_des_parametres(db, monkeypatch):
    monkeypatch.setattr(
        "apps.common.pdf.html_vers_pdf", lambda html, *, base_url=None: html.encode()
    )
    params = ParametresAssociation.load()
    params.signataire_nom = "Bureau Test"
    params.signataire_qualite = "Trésorier"
    params.save()
    recu = _emettre()  # sans signataire choisi
    html = pdf_de_recu(recu).decode()
    assert "Bureau Test" in html


# --- Bilan par catégorie ----------------------------------------------------


def _transaction(saison, categorie, type_flux, statut, montant):
    return Transaction.objects.create(
        saison=saison,
        categorie=categorie,
        type_flux=type_flux,
        statut=statut,
        libelle="x",
        montant=Decimal(montant),
        date=date(2026, 3, 1),
    )


def test_bilan_ventile_par_categorie_et_statut(db):
    saison = Saison.objects.create(nom="2025-2026")
    subventions = Categorie.objects.create(nom="Subventions")
    materiel = Categorie.objects.create(nom="Matériel")

    _transaction(
        saison, subventions, Transaction.TypeFlux.RECETTE, Transaction.Statut.PREVU, "1000"
    )
    _transaction(
        saison, subventions, Transaction.TypeFlux.RECETTE, Transaction.Statut.REALISE, "800"
    )
    _transaction(saison, materiel, Transaction.TypeFlux.DEPENSE, Transaction.Statut.PREVU, "300")
    _transaction(saison, materiel, Transaction.TypeFlux.DEPENSE, Transaction.Statut.REALISE, "250")

    bilan = bilan_par_categorie(saison)
    lignes = {ligne["categorie"]: ligne for ligne in bilan["lignes"]}

    assert lignes["Subventions"]["recette_prevu"] == Decimal("1000")
    assert lignes["Subventions"]["recette_realise"] == Decimal("800")
    assert lignes["Matériel"]["depense_prevu"] == Decimal("300")
    assert lignes["Matériel"]["solde_realise"] == Decimal("-250")

    totaux = bilan["totaux"]
    assert totaux["recette_realise"] == Decimal("800")
    assert totaux["depense_realise"] == Decimal("250")
    assert totaux["solde_realise"] == Decimal("550")  # 800 − 250
    assert totaux["solde_prevu"] == Decimal("700")  # 1000 − 300


def test_bilan_regroupe_les_sans_categorie(db):
    saison = Saison.objects.create(nom="2025-2026")
    _transaction(saison, None, Transaction.TypeFlux.RECETTE, Transaction.Statut.REALISE, "50")
    bilan = bilan_par_categorie(saison)
    assert bilan["lignes"][0]["categorie"] == "Sans catégorie"
    assert bilan["totaux"]["recette_realise"] == Decimal("50")


def test_bilan_ignore_les_autres_saisons(db):
    saison = Saison.objects.create(nom="2025-2026")
    autre = Saison.objects.create(nom="2024-2025")
    _transaction(autre, None, Transaction.TypeFlux.RECETTE, Transaction.Statut.REALISE, "999")
    bilan = bilan_par_categorie(saison)
    assert bilan["lignes"] == []
    assert bilan["totaux"]["recette_realise"] == Decimal("0.00")


def test_classeur_bilan_produit_un_xlsx_avec_les_donnees(db):
    from io import BytesIO

    from openpyxl import load_workbook

    saison = Saison.objects.create(nom="2025-2026")
    subventions = Categorie.objects.create(nom="Subventions")
    _transaction(
        saison, subventions, Transaction.TypeFlux.RECETTE, Transaction.Statut.REALISE, "800"
    )

    contenu = classeur_bilan(saison)
    assert contenu[:2] == b"PK"  # un .xlsx est une archive ZIP

    feuille = load_workbook(BytesIO(contenu)).active
    lignes = list(feuille.iter_rows(values_only=True))
    assert lignes[0][0] == "Catégorie"  # en-tête
    assert lignes[1][0] == "Subventions"
    assert lignes[1][2] == 800.0  # recettes réalisées
    assert lignes[-1][0] == "Total"  # ligne de totaux en dernier


def test_tresorerie_ajoute_le_reste_a_realiser_au_solde(db):
    saison = Saison.objects.create(nom="2025-2026")
    cat = Categorie.objects.create(nom="Général")
    solde = SoldeTresorerie.charger()
    solde.montant = Decimal("1000.00")
    solde.save()
    _transaction(saison, cat, Transaction.TypeFlux.RECETTE, Transaction.Statut.PREVU, "300")
    _transaction(saison, cat, Transaction.TypeFlux.RECETTE, Transaction.Statut.REALISE, "100")
    _transaction(saison, cat, Transaction.TypeFlux.DEPENSE, Transaction.Statut.PREVU, "120")

    t = tresorerie(saison)
    assert t["solde_pointe"] == Decimal("1000.00")
    assert t["reste_a_encaisser"] == Decimal("200.00")  # 300 prévu − 100 réalisé
    assert t["reste_a_decaisser"] == Decimal("120.00")  # 120 prévu − 0 réalisé
    assert t["previsionnelle"] == Decimal("1080.00")  # 1000 + 200 − 120


def test_tresorerie_reste_a_realiser_jamais_negatif(db):
    """Si le réalisé dépasse le prévu, le « reste à réaliser » est nul (pas
    négatif) : la prévisionnelle ne descend pas sous le solde pointé de ce fait."""
    saison = Saison.objects.create(nom="2025-2026")
    cat = Categorie.objects.create(nom="Général")
    solde = SoldeTresorerie.charger()
    solde.montant = Decimal("2000.00")
    solde.save()
    # Recettes réalisées (800) > prévues (500) ; dépenses réalisées (400) > prévues (300).
    _transaction(saison, cat, Transaction.TypeFlux.RECETTE, Transaction.Statut.PREVU, "500")
    _transaction(saison, cat, Transaction.TypeFlux.RECETTE, Transaction.Statut.REALISE, "800")
    _transaction(saison, cat, Transaction.TypeFlux.DEPENSE, Transaction.Statut.PREVU, "300")
    _transaction(saison, cat, Transaction.TypeFlux.DEPENSE, Transaction.Statut.REALISE, "400")

    t = tresorerie(saison)
    assert t["reste_a_encaisser"] == Decimal("0.00")
    assert t["reste_a_decaisser"] == Decimal("0.00")
    assert t["previsionnelle"] == Decimal("2000.00")  # rien à réaliser → = solde pointé


def test_tresorerie_sans_saison_ne_projette_pas(db):
    solde = SoldeTresorerie.charger()
    solde.montant = Decimal("500.00")
    solde.save()
    t = tresorerie(None)
    assert t["previsionnelle"] == Decimal("500.00")
    assert t["reste_a_encaisser"] == Decimal("0.00")


def test_solde_tresorerie_est_un_singleton(db):
    a = SoldeTresorerie.charger()
    a.montant = Decimal("100.00")
    a.save()
    b = SoldeTresorerie(montant=Decimal("999.00"))
    b.save()
    assert SoldeTresorerie.objects.count() == 1
    assert SoldeTresorerie.charger().montant == Decimal("999.00")


# --- Tableau de bord budgétaire (BUD-1) -------------------------------------
#
# `graphiques` ne calcule aucun montant : il met en forme la sortie de
# `bilan_par_categorie`. La plupart de ces tests lui passent donc un bilan
# construit à la main, sans base. Le dernier branche la VRAIE sortie du service
# pour que les deux formats ne puissent pas diverger en silence.


def _ligne_bilan(nom, *, rp="0", rr="0", dp="0", dr="0"):
    """Une ligne au format exact de `bilan_par_categorie`."""
    ligne = {
        "categorie": nom,
        "recette_prevu": Decimal(rp),
        "recette_realise": Decimal(rr),
        "depense_prevu": Decimal(dp),
        "depense_realise": Decimal(dr),
    }
    ligne["solde_prevu"] = ligne["recette_prevu"] - ligne["depense_prevu"]
    ligne["solde_realise"] = ligne["recette_realise"] - ligne["depense_realise"]
    return ligne


def _bilan(*lignes):
    totaux = _ligne_bilan("Total")
    for ligne in lignes:
        for cle in ("recette_prevu", "recette_realise", "depense_prevu", "depense_realise"):
            totaux[cle] += ligne[cle]
    totaux["solde_prevu"] = totaux["recette_prevu"] - totaux["depense_prevu"]
    totaux["solde_realise"] = totaux["recette_realise"] - totaux["depense_realise"]
    return {"lignes": list(lignes), "totaux": totaux}


_TRESO_VIDE = {"previsionnelle": Decimal("0.00"), "solde_pointe": Decimal("0.00")}


def test_tuiles_nomment_le_signe_du_solde_en_toutes_lettres():
    """La couleur seule ne doit jamais dire si l'exercice est excédentaire."""
    excedent = graphiques.tuiles(_bilan(_ligne_bilan("A", rr="900", dr="400")), _TRESO_VIDE)
    deficit = graphiques.tuiles(_bilan(_ligne_bilan("A", rr="100", dr="400")), _TRESO_VIDE)
    equilibre = graphiques.tuiles(_bilan(_ligne_bilan("A", rr="400", dr="400")), _TRESO_VIDE)

    solde_excedent, solde_deficit, solde_equilibre = (t[2] for t in (excedent, deficit, equilibre))
    assert solde_excedent["sens"] == "positif"
    assert "excédent" in solde_excedent["detail"]
    assert solde_deficit["sens"] == "negatif"
    assert "déficit" in solde_deficit["detail"]
    assert solde_equilibre["sens"] == "nul"
    assert "équilibre" in solde_equilibre["detail"]


def test_tuiles_laissent_les_montants_en_decimal():
    """La mise en forme (virgule française) appartient au gabarit, pas ici."""
    tuiles = graphiques.tuiles(_bilan(_ligne_bilan("A", rr="900", rp="1000")), _TRESO_VIDE)
    assert tuiles[0]["valeur"] == Decimal("900")
    assert tuiles[0]["detail_montant"] == Decimal("1000")


def test_comparaison_partage_une_seule_echelle_entre_les_deux_flux():
    """Une dépense de 500 € ne doit pas paraître aussi longue qu'une recette
    de 1 000 € : les deux graphiques se lisent l'un sous l'autre."""
    bilan = _bilan(_ligne_bilan("Général", rr="1000", dr="500"))
    comparaison = graphiques.comparaison_au_budget(bilan)

    assert comparaison["echelle"] == Decimal("1000")
    par_titre = {f["titre"]: f for f in comparaison["flux"]}
    assert par_titre["Recettes"]["lignes"][0]["part_realise"] == "100"
    assert par_titre["Dépenses"]["lignes"][0]["part_realise"] == "50.0"


def test_comparaison_ecarte_les_categories_sans_montant_sur_le_flux():
    """Une catégorie sans recette n'occupe pas une ligne vide du graphique
    des recettes — elle reste dans celui des dépenses."""
    bilan = _bilan(
        _ligne_bilan("Subventions", rr="800"),
        _ligne_bilan("Matériel", dr="300"),
    )
    par_titre = {f["titre"]: f for f in graphiques.comparaison_au_budget(bilan)["flux"]}

    assert [ligne["categorie"] for ligne in par_titre["Recettes"]["lignes"]] == ["Subventions"]
    assert [ligne["categorie"] for ligne in par_titre["Dépenses"]["lignes"]] == ["Matériel"]


def test_comparaison_classe_les_plus_gros_montants_en_premier():
    bilan = _bilan(
        _ligne_bilan("Petite", dr="100"),
        _ligne_bilan("Grosse", dr="900"),
        _ligne_bilan("Moyenne", dr="400"),
    )
    lignes = graphiques.comparaison_au_budget(bilan)["flux"][0]["lignes"]
    assert [ligne["categorie"] for ligne in lignes] == ["Grosse", "Moyenne", "Petite"]


def test_comparaison_qualifie_l_ecart_au_budget_par_un_mot():
    bilan = _bilan(
        _ligne_bilan("Dépassée", dp="100", dr="250"),
        _ligne_bilan("Économe", dp="500", dr="200"),
        _ligne_bilan("Juste", dp="300", dr="300"),
    )
    lignes = {
        ligne["categorie"]: ligne
        for ligne in graphiques.comparaison_au_budget(bilan)["flux"][0]["lignes"]
    }

    assert lignes["Dépassée"]["sens"] == "au-dessus"
    assert lignes["Dépassée"]["ecart_absolu"] == Decimal("150")
    assert lignes["Économe"]["sens"] == "en dessous"
    assert lignes["Économe"]["ecart_absolu"] == Decimal("300")  # valeur absolue, pas −300
    assert lignes["Juste"]["sens"] == "conforme"


def test_comparaison_sur_un_budget_vide_ne_divise_pas_par_zero():
    assert graphiques.comparaison_au_budget(_bilan())["flux"] == []


def test_repartition_regroupe_la_traine_au_dela_du_maximum():
    """Au-delà de sept parts, les teintes cessent d'être distinguables."""
    bilan = _bilan(*[_ligne_bilan(f"Cat {i}", dr=str(100 - i)) for i in range(10)])
    repartition = graphiques.repartition_depenses(bilan)

    assert len(repartition["segments"]) == graphiques.MAX_SEGMENTS
    dernier = repartition["segments"][-1]
    assert dernier["categorie"] == graphiques.ETIQUETTE_AUTRES
    assert repartition["regroupees"] == 4  # 10 catégories − 6 nommées
    # Rien n'est perdu en route : la traîne est additionnée, pas coupée.
    assert sum(s["montant"] for s in repartition["segments"]) == repartition["total"]


def test_repartition_ne_regroupe_rien_quand_le_compte_tient():
    bilan = _bilan(*[_ligne_bilan(f"Cat {i}", dr="100") for i in range(graphiques.MAX_SEGMENTS)])
    repartition = graphiques.repartition_depenses(bilan)

    assert repartition["regroupees"] == 0
    assert graphiques.ETIQUETTE_AUTRES not in [s["categorie"] for s in repartition["segments"]]


def test_repartition_ecrit_la_part_avec_un_point_pour_la_feuille_de_style():
    """Le projet tourne en fr-fr : une part rendue « 12,5 » casserait la
    largeur CSS dans laquelle elle est injectée. Les deux formes coexistent."""
    bilan = _bilan(_ligne_bilan("A", dr="125"), _ligne_bilan("B", dr="875"))
    segments = {s["categorie"]: s for s in graphiques.repartition_depenses(bilan)["segments"]}

    assert segments["A"]["part"] == "12.5"
    assert segments["A"]["part_lisible"] == "12,5"


def test_repartition_ignore_les_depenses_nulles_et_le_total_vide():
    vide = graphiques.repartition_depenses(_bilan(_ligne_bilan("A", rr="500")))
    assert vide["segments"] == []
    assert vide["total"] == Decimal("0.00")


def test_le_tableau_de_bord_consomme_la_vraie_sortie_du_bilan(db):
    """Garde-fou anti-dérive : si `bilan_par_categorie` change de format, ce
    test tombe avant l'écran."""
    saison = Saison.objects.create(nom="2025-2026")
    subventions = Categorie.objects.create(nom="Subventions")
    materiel = Categorie.objects.create(nom="Matériel")
    _transaction(
        saison, subventions, Transaction.TypeFlux.RECETTE, Transaction.Statut.REALISE, "800"
    )
    _transaction(saison, materiel, Transaction.TypeFlux.DEPENSE, Transaction.Statut.PREVU, "300")
    _transaction(saison, materiel, Transaction.TypeFlux.DEPENSE, Transaction.Statut.REALISE, "250")

    bilan = bilan_par_categorie(saison)
    tuiles = graphiques.tuiles(bilan, tresorerie(saison))
    comparaison = graphiques.comparaison_au_budget(bilan)
    repartition = graphiques.repartition_depenses(bilan)

    assert tuiles[0]["valeur"] == Decimal("800")  # recettes réalisées
    assert comparaison["echelle"] == Decimal("800")
    assert {f["titre"] for f in comparaison["flux"]} == {"Recettes", "Dépenses"}
    assert repartition["segments"][0]["categorie"] == "Matériel"
    assert repartition["segments"][0]["part"] == "100"


def test_repartition_borne_la_part_des_montants_negatifs():
    """Le formulaire du back-office refuse les montants négatifs, mais l'admin
    Django ne passe pas par lui. Une part négative produirait `width: -6.3%` —
    ignorée en silence par le navigateur."""
    bilan = _bilan(_ligne_bilan("Remboursement", dr="-500"), _ligne_bilan("Salaires", dr="8000"))
    lignes = {
        ligne["categorie"]: ligne
        for ligne in graphiques.comparaison_au_budget(bilan)["flux"][0]["lignes"]
    }

    assert lignes["Remboursement"]["part_realise"] == "0"
    # Le montant réel n'est pas escamoté : il reste écrit à côté de la barre.
    assert lignes["Remboursement"]["realise"] == Decimal("-500")


def test_repartition_ne_depasse_jamais_le_nombre_de_teintes_disponibles():
    """`maximum` ne peut que réduire le plafond : au-delà, les rangs n'ont plus
    de teinte en face côté feuille de style."""
    bilan = _bilan(*[_ligne_bilan(f"Cat {i}", dr="100") for i in range(20)])
    segments = graphiques.repartition_depenses(bilan, maximum=15)["segments"]

    assert len(segments) == graphiques.MAX_SEGMENTS
    assert max(s["rang"] for s in segments) == graphiques.MAX_SEGMENTS


def test_le_regroupement_met_toujours_au_moins_deux_categories_de_cote():
    """Invariant qui dispense le gabarit d'un cas singulier : regrouper la
    traîne suppose d'avoir dépassé le plafond, donc d'en mettre deux de côté."""
    for total in range(1, 14):
        for maximum in range(2, graphiques.MAX_SEGMENTS + 1):
            bilan = _bilan(*[_ligne_bilan(f"Cat {i}", dr="100") for i in range(total)])
            regroupees = graphiques.repartition_depenses(bilan, maximum=maximum)["regroupees"]
            assert regroupees == 0 or regroupees >= 2, (total, maximum, regroupees)
